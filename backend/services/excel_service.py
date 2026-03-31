"""
LegalEase AI - Excel Export Service
Builds a premium multi-sheet Excel workbook for founder-facing reports.
"""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


NAVY = "0B1220"
SLATE = "1E293B"
GOLD = "C28A2E"
GOLD_SOFT = "F5E6C8"
IVORY = "FAF7F2"
WHITE = "FFFFFF"
TEXT = "18212F"
TEXT_MUTED = "5B6472"
GREEN = "1F7A4C"
AMBER = "B76E12"
RED = "A13C32"
BLUE = "2B5C9A"
LINE = "D9D1C6"

THIN = Side(style="thin", color=LINE)
BOX_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FONT_HERO = Font(name="Aptos Display", size=20, bold=True, color=WHITE)
FONT_TITLE = Font(name="Aptos Display", size=14, bold=True, color=TEXT)
FONT_SECTION = Font(name="Aptos", size=11, bold=True, color=WHITE)
FONT_LABEL = Font(name="Aptos", size=9, bold=True, color=TEXT_MUTED)
FONT_BODY = Font(name="Aptos", size=10, color=TEXT)
FONT_BODY_BOLD = Font(name="Aptos", size=10, bold=True, color=TEXT)
FONT_METRIC = Font(name="Aptos Display", size=18, bold=True, color=WHITE)
FONT_SMALL = Font(name="Aptos", size=9, color=TEXT_MUTED)

FILL_HERO = PatternFill("solid", fgColor=NAVY)
FILL_SECTION = PatternFill("solid", fgColor=SLATE)
FILL_CARD = PatternFill("solid", fgColor=WHITE)
FILL_CANVAS = PatternFill("solid", fgColor=IVORY)
FILL_GOLD = PatternFill("solid", fgColor=GOLD)
FILL_GREEN = PatternFill("solid", fgColor=GREEN)
FILL_AMBER = PatternFill("solid", fgColor=AMBER)
FILL_RED = PatternFill("solid", fgColor=RED)
FILL_BLUE = PatternFill("solid", fgColor=BLUE)
FILL_SOFT_GOLD = PatternFill("solid", fgColor=GOLD_SOFT)


def _currency_text(value: str) -> str:
    return (value or "").replace("â‚¹", "Rs. ").replace("₹", "Rs. ")


def _safe_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def _set_background(ws, cols: int = 12, rows: int = 80) -> None:
    for row in range(1, rows + 1):
        for col in range(1, cols + 1):
            ws.cell(row, col).fill = FILL_CANVAS


def _hide_grid(ws) -> None:
    ws.sheet_view.showGridLines = False


def _hide_unused(ws, used_cols: int, used_rows: int) -> None:
    last_visible_col = min(used_cols + 14, 52)
    if last_visible_col > used_cols:
        ws.column_dimensions.group(get_column_letter(used_cols + 1), get_column_letter(last_visible_col), hidden=True)
    last_visible_row = min(used_rows + 80, 220)
    if last_visible_row > used_rows:
        ws.row_dimensions.group(used_rows + 1, last_visible_row, hidden=True)


def _set_col_widths(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _merge_fill(ws, cell_range: str, fill: PatternFill, border: Border = BOX_BORDER) -> None:
    ws.merge_cells(cell_range)
    start = ws[cell_range.split(":")[0]]
    start.fill = fill
    start.border = border
    for row in ws[cell_range]:
        for cell in row:
            cell.fill = fill
            cell.border = border


def _write(ws, cell: str, value: Any, font: Font = FONT_BODY, fill: PatternFill | None = None,
           alignment: Alignment | None = None, border: Border | None = None) -> None:
    target = ws[cell]
    target.value = value
    target.font = font
    if fill:
        target.fill = fill
    if alignment:
        target.alignment = alignment
    if border:
        target.border = border


def _card(ws, col1: str, col2: str, row_top: int, title: str, value: Any, note: str, fill: PatternFill) -> None:
    _merge_fill(ws, f"{col1}{row_top}:{col2}{row_top}", fill)
    _merge_fill(ws, f"{col1}{row_top+1}:{col2}{row_top+2}", fill)
    _merge_fill(ws, f"{col1}{row_top+3}:{col2}{row_top+3}", fill)
    _write(ws, f"{col1}{row_top}", title, font=FONT_SECTION, fill=fill, alignment=Alignment(horizontal="left", vertical="center"))
    _write(ws, f"{col1}{row_top+1}", value, font=FONT_METRIC, fill=fill, alignment=Alignment(horizontal="left", vertical="center"))
    _write(ws, f"{col1}{row_top+3}", note, font=Font(name="Aptos", size=9, color=WHITE), fill=fill, alignment=Alignment(wrap_text=True))


def _section_header(ws, row: int, title: str) -> None:
    _merge_fill(ws, f"A{row}:L{row}", FILL_SECTION)
    _write(ws, f"A{row}", title, font=FONT_SECTION, fill=FILL_SECTION, alignment=Alignment(horizontal="left", vertical="center"))


def _styled_table(ws, start_row: int, headers: list[str], rows: list[list[Any]], widths: list[int] | None = None) -> int:
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(start_row, idx, header)
        cell.font = FONT_SECTION
        cell.fill = FILL_SECTION
        cell.border = BOX_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row in enumerate(rows, start=start_row + 1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row_idx, col_idx, value)
            cell.font = FONT_BODY
            cell.border = BOX_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.fill = FILL_CARD if row_idx % 2 else FILL_CANVAS

    if widths:
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width

    return start_row + len(rows)


def _overview_sheet(wb: Workbook, data: dict[str, Any]) -> None:
    ws = wb.active
    ws.title = "Executive Dashboard"
    _hide_grid(ws)
    _set_background(ws, cols=12, rows=44)
    _set_col_widths(ws, {
        "A": 20, "B": 16, "C": 16, "D": 3, "E": 20, "F": 16, "G": 16, "H": 3,
        "I": 20, "J": 16, "K": 16, "L": 3,
    })
    ws.sheet_view.zoomScale = 90
    ws.freeze_panes = "A5"
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[13].height = 30
    ws.row_dimensions[14].height = 30
    ws.row_dimensions[16].height = 30

    _merge_fill(ws, "A1:L1", FILL_HERO, border=Border())
    _merge_fill(ws, "A2:L2", FILL_HERO, border=Border())
    _merge_fill(ws, "A3:L3", FILL_HERO, border=Border())
    _write(ws, "A1", f"LegalEase Executive Report", font=FONT_HERO, fill=FILL_HERO)
    _write(ws, "A2", data.get("business_name", "Business"), font=Font(name="Aptos", size=11, bold=True, color=GOLD_SOFT), fill=FILL_HERO)
    meta = (
        f"Report ID: {data.get('report_id', '-')}   |   Category: {data.get('category', '-').title()}   |   "
        f"Location: {data.get('location', '-')}   |   Scale: {data.get('scale', '-')}"
    )
    _write(ws, "A3", meta, font=Font(name="Aptos", size=9, color="D7DEEA"), fill=FILL_HERO)

    feasibility = data.get("feasibility", {})
    risk = data.get("risk", {})
    complexity = data.get("compliance_complexity", {})
    licenses = data.get("licenses", []) or []
    action_plan = data.get("action_plan", []) or []
    risks = data.get("risks", []) or []

    _card(ws, "A", "C", 5, "Feasibility", feasibility.get("score", 0), feasibility.get("label", ""), FILL_GREEN)
    _card(ws, "E", "G", 5, "Risk Exposure", risk.get("score", 0), risk.get("label", ""), FILL_RED if risk.get("score", 0) > 60 else FILL_AMBER)
    _card(ws, "I", "K", 5, "Compliance Load", complexity.get("score", 0), complexity.get("label", ""), FILL_BLUE)

    _section_header(ws, 11, "Founder Summary")
    _merge_fill(ws, "A12:G12", FILL_CARD)
    _merge_fill(ws, "A13:G14", FILL_CARD)
    _merge_fill(ws, "A15:G15", FILL_CARD)
    _merge_fill(ws, "A16:G16", FILL_CARD)
    _write(ws, "A12", "Executive Summary", font=FONT_TITLE, fill=FILL_CARD)
    _write(
        ws,
        "A13",
        data.get("summary", ""),
        fill=FILL_CARD,
        alignment=Alignment(wrap_text=True, vertical="top"),
    )
    _write(ws, "A15", "Key Insight", font=FONT_LABEL, fill=FILL_CARD)
    _write(
        ws,
        "A16",
        data.get("key_insight", ""),
        font=FONT_BODY_BOLD,
        fill=FILL_CARD,
        alignment=Alignment(wrap_text=True, vertical="top"),
    )

    for row_idx in range(12, 17):
        for col in range(9, 12):
            ws.cell(row_idx, col).fill = FILL_CARD
            ws.cell(row_idx, col).border = BOX_BORDER
            ws.cell(row_idx, col).alignment = Alignment(vertical="center")
    _write(ws, "I12", "At a Glance", font=FONT_TITLE, fill=FILL_CARD)
    glance_rows = [
        ("Licenses required", len(licenses)),
        ("High-priority items", sum(1 for lic in licenses if (lic.get("priority") or "").lower() in {"critical", "high"})),
        ("Legal risks flagged", len(risks)),
        ("Action steps", len(action_plan)),
    ]
    row = 13
    for label, value in glance_rows:
        _write(ws, f"I{row}", label, font=FONT_LABEL, fill=FILL_CARD)
        _write(ws, f"K{row}", value, font=FONT_BODY_BOLD, fill=FILL_CARD, alignment=Alignment(horizontal="right"))
        row += 1

    chart_start = 19
    _section_header(ws, chart_start, "Visual Snapshot")
    score_data_row = chart_start + 2
    ws[f"A{score_data_row}"] = "Metric"
    ws[f"B{score_data_row}"] = "Score"
    metrics = [
        ("Feasibility", feasibility.get("score", 0)),
        ("Risk", risk.get("score", 0)),
        ("Complexity", complexity.get("score", 0)),
    ]
    for idx, (label, value) in enumerate(metrics, start=1):
        ws[f"A{score_data_row + idx}"] = label
        ws[f"B{score_data_row + idx}"] = value

    priority_row = chart_start + 2
    ws[f"I{priority_row}"] = "Priority"
    ws[f"J{priority_row}"] = "Count"
    priority_counts = {"Critical": 0, "High": 0, "Medium": 0, "Low": 0}
    for lic in licenses:
        priority = (lic.get("priority") or "low").title()
        priority_counts[priority] = priority_counts.get(priority, 0) + 1
    for idx, (label, value) in enumerate(priority_counts.items(), start=1):
        ws[f"I{priority_row + idx}"] = label
        ws[f"J{priority_row + idx}"] = value

    for addr in [f"A{score_data_row}:B{score_data_row}", f"I{priority_row}:J{priority_row}"]:
        for row_cells in ws[addr]:
            for cell in row_cells:
                cell.fill = FILL_SECTION
                cell.font = FONT_SECTION
                cell.border = BOX_BORDER
                cell.alignment = Alignment(horizontal="center")

    bar = BarChart()
    bar.title = "Core Score Comparison"
    bar.y_axis.title = "Score"
    bar.style = 10
    bar.legend = None
    bar.height = 6.8
    bar.width = 8.8
    bar.gapWidth = 45
    bar.add_data(Reference(ws, min_col=2, min_row=score_data_row, max_row=score_data_row + 3), titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=1, min_row=score_data_row + 1, max_row=score_data_row + 3))
    ws.add_chart(bar, "D21")

    doughnut = DoughnutChart()
    doughnut.title = "License Priority Distribution"
    doughnut.height = 6.8
    doughnut.width = 6.8
    doughnut.holeSize = 58
    doughnut.varyColors = True
    doughnut.add_data(Reference(ws, min_col=10, min_row=priority_row, max_row=priority_row + 4), titles_from_data=True)
    doughnut.set_categories(Reference(ws, min_col=9, min_row=priority_row + 1, max_row=priority_row + 4))
    doughnut.dLbls = DataLabelList()
    doughnut.dLbls.showPercent = True
    doughnut.dLbls.showLeaderLines = True
    doughnut.legend = None
    ws.add_chart(doughnut, "I21")

    _section_header(ws, 33, "Immediate Actions")
    top_steps = action_plan[:5] if action_plan else []
    if not top_steps:
        top_steps = [{"step": 1, "title": "No action plan available", "timeframe": "-", "cost": "-", "description": "-"}]
    for idx, step in enumerate(top_steps, start=34):
        for col in range(1, 13):
            ws.cell(idx, col).fill = FILL_CARD
            ws.cell(idx, col).border = BOX_BORDER
        ws.row_dimensions[idx].height = 22
        _write(ws, f"A{idx}", f"Step {step.get('step', idx-33)}", font=Font(name="Aptos", size=9, bold=True, color=GOLD), fill=FILL_CARD)
        _write(ws, f"B{idx}", step.get("title", ""), font=FONT_BODY_BOLD, fill=FILL_CARD, alignment=Alignment(wrap_text=True))
        _write(ws, f"H{idx}", step.get("timeframe", ""), font=FONT_SMALL, fill=FILL_CARD)
        _write(ws, f"J{idx}", _currency_text(step.get("cost", "")), font=FONT_SMALL, fill=FILL_CARD)

    ws.print_area = "A1:L39"
    _hide_unused(ws, used_cols=12, used_rows=44)


def _detail_sheet(ws, title: str, subtitle: str) -> None:
    _hide_grid(ws)
    _set_background(ws, cols=7, rows=140)
    _merge_fill(ws, "A1:G1", FILL_HERO, border=Border())
    _merge_fill(ws, "A2:G2", FILL_HERO, border=Border())
    _write(ws, "A1", title, font=FONT_HERO, fill=FILL_HERO)
    _write(ws, "A2", subtitle, font=Font(name="Aptos", size=10, color="D7DEEA"), fill=FILL_HERO)
    ws.sheet_view.zoomScale = 90
    ws.freeze_panes = "A5"
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 20


def _licenses_sheet(wb: Workbook, data: dict[str, Any]) -> None:
    ws = wb.create_sheet("Licensing Roadmap")
    _detail_sheet(ws, "Licensing Roadmap", "All registrations, costs, priorities, and portal references.")
    rows = [
        [
            lic.get("name", ""),
            (lic.get("priority", "") or "").title(),
            _currency_text(lic.get("estimated_cost", "")),
            lic.get("time_to_approve", ""),
            lic.get("authority", ""),
            lic.get("description", ""),
            lic.get("link", ""),
        ]
        for lic in data.get("licenses", []) or []
    ]
    _styled_table(
        ws,
        5,
        ["License", "Priority", "Cost", "Approval Window", "Authority", "Why It Matters", "Portal"],
        rows,
        widths=[30, 14, 16, 18, 26, 52, 34],
    )
    for row in range(6, max(ws.max_row, 20) + 1):
        ws.row_dimensions[row].height = 36
    ws.print_area = f"A1:G{max(ws.max_row, 20)}"
    _hide_unused(ws, used_cols=7, used_rows=max(ws.max_row + 2, 24))


def _risks_sheet(wb: Workbook, data: dict[str, Any]) -> None:
    ws = wb.create_sheet("Risk Register")
    _detail_sheet(ws, "Risk Register", "Penalty exposure, severity, and governing law.")
    rows = [
        [
            risk.get("title", ""),
            (risk.get("severity", "") or "").title(),
            _currency_text(risk.get("penalty", "")),
            risk.get("law", ""),
            risk.get("description", ""),
        ]
        for risk in data.get("risks", []) or []
    ]
    _styled_table(ws, 5, ["Risk", "Severity", "Penalty", "Law", "Description"], rows, widths=[28, 14, 18, 28, 62])
    for row in range(6, max(ws.max_row, 18) + 1):
        ws.row_dimensions[row].height = 34

    breakdown_row = max(ws.max_row + 3, 14)
    _section_header(ws, breakdown_row, "Risk Breakdown")
    breakdown = data.get("risk_breakdown", []) or []
    for idx, item in enumerate(breakdown, start=breakdown_row + 1):
        ws[f"A{idx}"] = item.get("label", "")
        ws[f"B{idx}"] = item.get("score", 0)
        ws[f"A{idx}"].font = FONT_BODY
        ws[f"B{idx}"].font = FONT_BODY_BOLD
        ws[f"A{idx}"].fill = FILL_CARD
        ws[f"B{idx}"].fill = FILL_CARD
        ws[f"A{idx}"].border = BOX_BORDER
        ws[f"B{idx}"].border = BOX_BORDER

    if breakdown:
        chart = BarChart()
        chart.title = "Risk Breakdown by Dimension"
        chart.height = 7
        chart.width = 9
        chart.y_axis.title = "Score"
        chart.legend = None
        chart.add_data(Reference(ws, min_col=2, min_row=breakdown_row, max_row=breakdown_row + len(breakdown)), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=breakdown_row + 1, max_row=breakdown_row + len(breakdown)))
        ws.add_chart(chart, "F15")
    ws.print_area = f"A1:G{max(ws.max_row + 2, 26)}"
    _hide_unused(ws, used_cols=7, used_rows=max(ws.max_row + 2, 26))


def _action_sheet(wb: Workbook, data: dict[str, Any]) -> None:
    ws = wb.create_sheet("Execution Plan")
    _detail_sheet(ws, "Execution Plan", "Chronological operating plan with ownership-ready steps.")
    rows = [
        [
            step.get("step", ""),
            step.get("title", ""),
            (step.get("category", "") or "").title(),
            step.get("timeframe", ""),
            _currency_text(step.get("cost", "")),
            step.get("description", ""),
        ]
        for step in data.get("action_plan", []) or []
    ]
    _styled_table(ws, 5, ["Step", "Action", "Category", "Timeframe", "Cost", "Execution Detail"], rows, widths=[8, 28, 18, 18, 14, 60])
    for row in range(6, max(ws.max_row, 18) + 1):
        ws.row_dimensions[row].height = 38
    ws.print_area = f"A1:F{max(ws.max_row, 20)}"
    _hide_unused(ws, used_cols=6, used_rows=max(ws.max_row + 2, 22))


def _costs_sheet(wb: Workbook, data: dict[str, Any]) -> None:
    ws = wb.create_sheet("Cost Planner")
    _detail_sheet(ws, "Cost Planner", "Cost ranges, planning assumptions, and founder budgeting notes.")
    rows = [
        [
            cost.get("item", ""),
            _currency_text(cost.get("range", "")),
            cost.get("notes", ""),
        ]
        for cost in data.get("cost_estimates", []) or []
    ]
    _styled_table(ws, 5, ["Cost Item", "Estimated Range", "Planning Note"], rows, widths=[30, 20, 68])
    for row in range(6, max(ws.max_row, 16) + 1):
        ws.row_dimensions[row].height = 34
    ws.print_area = f"A1:C{max(ws.max_row, 18)}"
    _hide_unused(ws, used_cols=3, used_rows=max(ws.max_row + 2, 20))


def _readiness_sheet(wb: Workbook, data: dict[str, Any]) -> None:
    ws = wb.create_sheet("Founder Brief")
    _detail_sheet(ws, "Founder Brief", "Condensed business context, compliance outlook, and next-check questions.")

    _section_header(ws, 5, "Business Profile")
    profile_rows = [
        ("Business Name", data.get("business_name", "")),
        ("Category", data.get("category", "").title()),
        ("Location", data.get("location", "")),
        ("Scale", data.get("scale", "")),
        ("Mode", data.get("mode", "")),
        ("Report Link", data.get("report_url", "")),
    ]
    row = 6
    for label, value in profile_rows:
        _merge_fill(ws, f"A{row}:B{row}", FILL_CARD)
        _merge_fill(ws, f"C{row}:G{row}", FILL_CARD)
        _write(ws, f"A{row}", label, font=FONT_LABEL, fill=FILL_CARD)
        _write(ws, f"C{row}", value, font=FONT_BODY_BOLD, fill=FILL_CARD, alignment=Alignment(wrap_text=True))
        ws.row_dimensions[row].height = 24
        row += 1

    _section_header(ws, 13, "Questions to Clarify")
    questions = data.get("follow_up_questions", []) or []
    if not questions:
        questions = ["No follow-up questions available."]
    for idx, question in enumerate(questions, start=14):
        _merge_fill(ws, f"A{idx}:G{idx}", FILL_CARD)
        _write(ws, f"A{idx}", f"{idx-13}. {question}", font=FONT_BODY, fill=FILL_CARD, alignment=Alignment(wrap_text=True))
        ws.row_dimensions[idx].height = 32
    ws.print_area = f"A1:G{max(ws.max_row, 22)}"
    _hide_unused(ws, used_cols=7, used_rows=max(ws.max_row + 2, 24))


def build_excel_report(data: dict[str, Any]) -> bytes:
    wb = Workbook()
    _overview_sheet(wb, data)
    _licenses_sheet(wb, data)
    _risks_sheet(wb, data)
    _action_sheet(wb, data)
    _costs_sheet(wb, data)
    _readiness_sheet(wb, data)

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()
